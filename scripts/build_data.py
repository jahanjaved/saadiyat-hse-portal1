
from pathlib import Path
from datetime import datetime
import json
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "source" / "Saadiyat_Lagoons_HSE_Full_System_Updated.xlsx"
OUT = ROOT / "data" / "site-data.json"
OVERRIDES = ROOT / "data" / "status-overrides.json"
EXAMPLE = ROOT / "data" / "status-overrides.example.json"

def fmt_date(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return v

def load_overrides():
    file = OVERRIDES if OVERRIDES.exists() else EXAMPLE
    if file.exists():
        try:
            return json.loads(file.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}

wbv = openpyxl.load_workbook(SOURCE, data_only=True)

# clusters
ws = wbv["Cluster_Master"]
cluster_headers = [ws.cell(2, c).value for c in range(1, 9)]
clusters = []
for r in range(3, ws.max_row + 1):
    if ws.cell(r, 1).value is None:
        continue
    rec = {}
    for i, h in enumerate(cluster_headers, start=1):
        rec[h] = ws.cell(r, i).value
    rec["key"] = f'{rec["Package"]}|{rec["Cluster"]}'
    clusters.append(rec)
cluster_lookup = {c["key"]: c for c in clusters}

# KPI weights
ws = wbv["KPI_Weights"]
kpi_weights = []
for r in range(3, ws.max_row + 1):
    if ws.cell(r, 1).value is None:
        continue
    kpi_weights.append({
        "kpi": ws.cell(r, 1).value,
        "weight": ws.cell(r, 2).value,
        "critical_gate": ws.cell(r, 3).value,
        "notes": ws.cell(r, 4).value,
    })

# inspections
wsv = wbv["Weekly_Inspections"]
ins_headers = [wsv.cell(3, c).value for c in range(1, 59)]
score_map = {
    "Work_at_Height": "Work at Height",
    "Edge_Protection": "Edge Protection",
    "Falling_Object_Prevention": "Falling Object Prevention",
    "Excavation_Safety": "Excavation Safety",
    "Scaffolding_Compliance": "Scaffolding Compliance",
    "PTW_Implementation": "PTW Implementation",
    "PTW_Field_Verification": "PTW Field Verification",
    "MSRA_Quality": "MSRA Quality",
    "Lifting/Precast Installation": "Lifting / Precast",
    "Traffic_Interface": "Traffic Interface",
    "Housekeeping/Waste Management": "Housekeeping / Waste",
    "Welfare Arrangement": "Welfare Arrangement",
    "Fire_Readiness": "Fire Readiness",
    "Supervision_Subcontractor": "Supervision / Subcontractor",
    "Electrical_Tool_Safety": "Electrical / Tool Safety",
}
inspections = []
for r in range(4, wsv.max_row + 1):
    if wsv.cell(r, 2).value is None and wsv.cell(r, 6).value is None:
        continue
    row = {ins_headers[c-1]: wsv.cell(r, c).value for c in range(1, 59)}
    package = row.get("Package")
    cluster = row.get("Cluster")
    lookup = cluster_lookup.get(f"{package}|{cluster}", {})
    ins_id = f"INS-{r-3:03d}"
    scores = {}
    for key, label in score_map.items():
        val = row.get(key)
        if isinstance(val, (int, float)):
            scores[label] = float(val)
    avg_score = round(sum(scores.values()) / len(scores), 2) if scores else None
    status = "Good"
    if row.get("Stop_Work") == "YES":
        status = "Critical"
    elif avg_score is not None and avg_score < 3:
        status = "Action Required"
    elif avg_score is not None and avg_score < 4:
        status = "Monitor"
    photos = []
    evidence = row.get("Evidence_Link_or_Photo")
    if evidence:
        text = str(evidence)
        if text.lower().endswith((".png",".jpg",".jpeg",".webp",".gif")):
            photos.append({"type": "image", "src": text, "caption": "Evidence image"})
        else:
            photos.append({"type": "link", "src": text, "caption": "Evidence link"})
    inspections.append({
        "inspection_id": ins_id,
        "inspection_date": fmt_date(row.get("Inspection_Date")),
        "week_no": row.get("Week_No"),
        "month": row.get("Month"),
        "package": package,
        "cluster": cluster,
        "contractor": lookup.get("Contractor") or row.get("Contractor"),
        "project_director": lookup.get("Project Director (PD)") or row.get("Project Director (PD)"),
        "project_manager": lookup.get("Project Manager (PM)") or row.get("Project Manager (PM)"),
        "construction_manager": lookup.get("Construction Manager (CM)") or row.get("Construction Manager (CM)"),
        "hse_manager": lookup.get("HSE Manager (HSEM)") or row.get("HSE Manager (HSEM)"),
        "section_engineer": lookup.get("Section Engineer (SE)") or row.get("Section Engineer (SE)"),
        "your_name": row.get("Your_Name"),
        "your_role": row.get("Your_Role"),
        "area_or_villa": row.get("Area_or_Villa"),
        "workers_observed": row.get("Workers_Observed"),
        "main_high_risk_activity": row.get("Main_High_Risk_Activity"),
        "stop_work": row.get("Stop_Work") == "YES",
        "scores": scores,
        "average_score": avg_score,
        "weighted_score": row.get("Weighted_Score_%"),
        "rating_band": row.get("Rating_Band"),
        "critical_red_flag": row.get("Critical_Red_Flag"),
        "award_eligible": row.get("Award_Eligible"),
        "repeat_issue_seen": row.get("Repeat_Issue_Seen"),
        "likely_root_cause": row.get("Likely_Root_Cause"),
        "top_3_gaps_observed": row.get("Top_3_Gaps_Observed"),
        "immediate_action_taken": row.get("Immediate_Action_Taken"),
        "preventive_action_required": row.get("Preventive_Action_Required"),
        "positive_observations": row.get("Positive_Observations"),
        "action_owner": row.get("Action_Owner"),
        "target_closeout_date": fmt_date(row.get("Target_Closeout_Date")),
        "status": status,
        "photos": photos,
    })

# actions
overrides = load_overrides()
wsc = wbv["CAPA_Tracker"]
actions = []
for r in range(4, wsc.max_row + 1):
    date_raised = wsc.cell(r, 2).value
    cluster = wsc.cell(r, 4).value
    if date_raised is None and cluster is None:
        continue
    helper_idx = wsc.cell(r, 20).value
    action_id = f"CAPA-{helper_idx:03d}" if isinstance(helper_idx, (int, float)) else f"CAPA-{r-3:03d}"
    ov = overrides.get(action_id, {})
    actions.append({
        "action_id": action_id,
        "date_raised": fmt_date(date_raised),
        "package": wsc.cell(r, 3).value,
        "cluster": cluster,
        "contractor": wsc.cell(r, 5).value,
        "kpi_area": wsc.cell(r, 6).value,
        "finding": wsc.cell(r, 7).value,
        "root_cause": wsc.cell(r, 8).value,
        "immediate_action": wsc.cell(r, 9).value,
        "preventive_action": wsc.cell(r, 10).value,
        "action_owner": wsc.cell(r, 11).value,
        "target_date": fmt_date(wsc.cell(r, 12).value),
        "status": ov.get("status", wsc.cell(r, 13).value or "Open"),
        "close_date": ov.get("close_date", fmt_date(wsc.cell(r, 14).value)),
        "days_open": wsc.cell(r, 15).value,
        "overdue_flag": wsc.cell(r, 16).value,
        "evidence_link": wsc.cell(r, 17).value,
        "comments": ov.get("comments", ""),
    })

avg_scores = [i["average_score"] for i in inspections if i["average_score"] is not None]
dashboard = {
    "site_title": "Saadiyat Lagoons HSE Action Portal",
    "generated_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
    "total_inspections": len(inspections),
    "total_actions": len(actions),
    "stop_work_cases": len([i for i in inspections if i["stop_work"]]),
    "repeat_issues": len([i for i in inspections if (i.get("repeat_issue_seen") or "").upper() == "YES"]),
    "avg_score": round(sum(avg_scores)/len(avg_scores), 2) if avg_scores else 0,
    "packages": sorted({i["package"] for i in inspections if i["package"] is not None}),
    "clusters_with_records": sorted({i["cluster"] for i in inspections if i["cluster"]}),
    "weeks": sorted({i["week_no"] for i in inspections if i["week_no"] is not None}),
    "months": sorted({i["month"] for i in inspections if i["month"] is not None}),
}
site_data = {
    "dashboard": dashboard,
    "kpi_weights": kpi_weights,
    "clusters": clusters,
    "inspections": inspections,
    "actions": actions,
}
OUT.write_text(json.dumps(site_data, indent=2), encoding="utf-8")
print(f"Built {OUT}")
