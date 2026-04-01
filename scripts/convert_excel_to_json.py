from pathlib import Path
import openpyxl, json
from datetime import datetime, date, time

SOURCE = Path('source/workbook.xlsx')
OUTPUT = Path('data/workbook-data.json')

def norm(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d %H:%M:%S') if v.time() != time(0,0) else v.strftime('%Y-%m-%d')
    if isinstance(v, date):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, time):
        return v.strftime('%H:%M:%S')
    return v

def parse_notes_sheet(ws):
    notes = []
    for r in range(1, ws.max_row + 1):
        vals = [norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        vals = [str(v).strip() for v in vals if v is not None and str(v).strip()]
        if vals:
            notes.append(' — '.join(vals))
    return {'type': 'notes', 'headers': [], 'rows': [], 'notes': notes}

def parse_table(ws, header_row, start_row=None):
    if start_row is None:
        start_row = header_row + 1
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h).strip() if h not in (None, '') else f'Column_{i}' for i, h in enumerate(headers, 1)]
    rows = []
    for r in range(start_row, ws.max_row + 1):
        vals = [norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if not any(str(v).strip() for v in vals):
            continue
        rows.append({headers[i]: vals[i] for i in range(len(headers))})
    return {'type': 'table', 'headers': headers, 'rows': rows, 'notes': []}

def parse_dashboard(ws, header_row, data_start_row, note_rows):
    notes = []
    for r in note_rows:
        vals = [norm(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 4) + 1)]
        if any(str(v).strip() for v in vals):
            notes.append(vals)
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h).strip() if h not in (None, '') else f'Column_{i}' for i, h in enumerate(headers, 1)]
    rows = []
    for r in range(data_start_row, ws.max_row + 1):
        vals = [norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if not any(str(v).strip() for v in vals):
            continue
        rows.append({headers[i]: vals[i] for i in range(len(headers))})
    return {'type': 'table', 'headers': headers, 'rows': rows, 'notes': notes}

def build():
    wb = openpyxl.load_workbook(SOURCE, data_only=True)
    sheet_order = wb.sheetnames
    sheets = {}

    for name in sheet_order:
        ws = wb[name]
        if name == 'Instructions':
            sheets[name] = parse_notes_sheet(ws)
        elif name in ('Cluster_Master', 'Lookups', 'KPI_Weights', 'Weekly_Inspections', 'CAPA_Tracker'):
            sheets[name] = parse_table(ws, 3, 4)
        elif name == 'Weekly_Dashboard':
            sheets[name] = parse_dashboard(ws, 9, 10, [1,2,4,5,6,7])
        elif name in ('Monthly_Dashboard', 'Gap_Analysis'):
            sheets[name] = parse_dashboard(ws, 7, 8, [1,2,4,5])
        else:
            header_row = None
            for r in range(1, min(ws.max_row, 15) + 1):
                vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                if sum(1 for v in vals if v not in (None, '')) >= 2:
                    header_row = r
                    break
            if header_row is None:
                sheets[name] = {'type': 'notes', 'headers': [], 'rows': [], 'notes': []}
            else:
                sheets[name] = parse_table(ws, header_row, header_row + 1)

    if 'CAPA_Tracker' in sheets:
        clean_capa = []
        for row in sheets['CAPA_Tracker']['rows']:
            meaningful = any(str(v).strip() not in ('', '0', '00:00:00') for v in row.values())
            if meaningful:
                clean_capa.append(row)
        sheets['CAPA_Tracker']['rows'] = clean_capa

    summary = {
        'workbook_name': SOURCE.name,
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'sheet_count': len(sheet_order),
        'sheet_names': sheet_order,
        'inspection_count': len(sheets.get('Weekly_Inspections', {}).get('rows', [])),
        'action_count': len(sheets.get('CAPA_Tracker', {}).get('rows', [])),
        'cluster_count': len(sheets.get('Cluster_Master', {}).get('rows', [])),
    }

    data = {'summary': summary, 'sheet_order': sheet_order, 'sheets': sheets}
    OUTPUT.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding='utf-8')
    print(f'Created {OUTPUT}')

if __name__ == '__main__':
    build()
